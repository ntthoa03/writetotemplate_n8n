from __future__ import annotations
import uvicorn
import glob
import os
import subprocess
import sys
import zipfile
import ctypes
import traceback
import threading
import time
import tempfile
from copy import copy
from shutil import which
from fastapi import BackgroundTasks
import shutil
import tempfile
from datetime import datetime, timedelta
import uuid
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel
from fastapi import APIRouter, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List
import tempfile
import shutil
import os
import base64

import os

API_KEY = os.getenv("API_KEY")
app = FastAPI()
JOBS: dict[str, dict] = {}

from fastapi import Header, HTTPException, Depends
 

def verify_api_key(x_api_key: str = Header(None)):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")
# ==================================================
# 1. TỰ ĐỘNG MỞ FIREWALL (CHỈ CHẠY TRÊN WINDOWS)
# ==================================================
def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

def open_firewall_port(port):
    if sys.platform == "win32":
        try:
            # Lệnh netsh để mở port trên Windows Firewall
            rule_name = f"Allow_n8n_Excel_Merge_{port}"
            # Kiểm tra xem rule đã tồn tại chưa để tránh add trùng
            check_cmd = f'netsh advfirewall firewall show rule name="{rule_name}"'
            result = subprocess.run(check_cmd, shell=True, capture_output=True, text=True)
            
            if "no rules match" in result.stdout.lower() or result.returncode != 0:
                print(f"[*] Dang tu dong mo Port {port} tren Firewall...")
                add_cmd = f'netsh advfirewall firewall add rule name="{rule_name}" dir=in action=allow protocol=TCP localport={port}'
                subprocess.run(add_cmd, shell=True, check=True, capture_output=True)
                print(f"[+] Da mo Port {port} thanh cong.")
            else:
                print(f"[OK] Port {port} da duoc mo truoc do.")
        except Exception as e:
            print(f"[!] Khong the mo Firewall: {e}")
            print("[!] Hay chay file .exe bang quyen 'Run as Administrator' neu n8n bi loi ket noi.")

# ==================================================
# ==================================================
# 1. KHAI BÁO BIẾN HỨNG DỮ LIỆU TỪ N8N (Thêm template_filename)
# ==================================================
class MergeRequest(BaseModel):
    folder_path: str
    template_filename: str  # Biến mới để nhận tên file mẫu
    run_in_background: bool = True

class PathDebugRequest(BaseModel):
    folder_path: str

# --- CÁC HÀM HELPER CỦA BẠN (GIỮ NGUYÊN) ---
def excel_date_to_str(serial):
    try:
        if pd.isna(serial) or str(serial).strip() == "": return ""
        date = datetime(1899, 12, 30) + timedelta(days=float(serial))
        return date.strftime("%d/%m/%Y")
    except Exception: return serial

def remove_footer(data_part):
    footer_keywords = ["Người lập", "Kiểm soát", "Created date", "Trang"]
    clean_rows = []
    for _, row in data_part.iterrows():
        row_text = " ".join([str(x) for x in row if pd.notna(x)])
        if any(k in row_text for k in footer_keywords): break
        clean_rows.append(row)
    df_clean = pd.DataFrame(clean_rows)
    df_clean = df_clean.dropna(how="all")
    return df_clean

def find_start_row(ws):
    start_row = 9
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == "STT": return cell.row + 1
    return start_row

def is_real_xlsx(path: str) -> bool:
    try:
        if not os.path.isfile(path) or os.path.getsize(path) < 64: return False
        return zipfile.is_zipfile(path)
    except OSError: return False

def _safe_sheet_title(name: str, index: int) -> str:
    bad = "[]:*?/\\"
    s = (name or f"Sheet{index + 1}")[:31]
    for c in bad: s = s.replace(c, "_")
    return s or f"Sheet{index + 1}"

def convert_xls_to_xlsx_xlrd_openpyxl(xls_path: str, out_xlsx: str) -> bool:
    try:
        import xlrd
        from xlrd import XL_CELL_BLANK, XL_CELL_BOOLEAN, XL_CELL_DATE, XL_CELL_EMPTY, XL_CELL_ERROR, XL_CELL_NUMBER, XL_CELL_TEXT
        from xlrd.xldate import xldate_as_datetime
        from openpyxl import Workbook
    except ImportError: return False
    xls_path = os.path.abspath(xls_path)
    out_xlsx = os.path.abspath(out_xlsx)
    def cell_value(book, cell):
        ctype = cell.ctype
        val = cell.value
        if ctype in (XL_CELL_EMPTY, XL_CELL_BLANK): return None
        if ctype == XL_CELL_TEXT: return val
        if ctype == XL_CELL_NUMBER: return val
        if ctype == XL_CELL_BOOLEAN: return bool(val)
        if ctype == XL_CELL_DATE:
            try: return xldate_as_datetime(val, book.datemode)
            except Exception: return val
        if ctype == XL_CELL_ERROR: return None
        return val
    try:
        book_xls = xlrd.open_workbook(xls_path, formatting_info=False)
        if book_xls.nsheets < 1: return False
        wb_out = Workbook()
        default_ws = wb_out.active
        first = True
        for si in range(book_xls.nsheets):
            sh = book_xls.sheet_by_index(si)
            if first:
                ws = default_ws
                ws.title = _safe_sheet_title(sh.name, si)
                first = False
            else: ws = wb_out.create_sheet(title=_safe_sheet_title(sh.name, si))
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    v = cell_value(book_xls, cell)
                    if v is not None and v != "": ws.cell(row=r + 1, column=c + 1, value=v)
            for mc in getattr(sh, "merged_cells", None) or []:
                try:
                    if isinstance(mc, tuple) and len(mc) == 4: rlo, rhi, clo, chi = int(mc[0]), int(mc[1]), int(mc[2]), int(mc[3])
                    elif hasattr(mc, "row_lo"): rlo, rhi, clo, chi = (int(mc.row_lo), int(mc.row_hi), int(mc.col_lo), int(mc.col_hi))
                    else: continue
                    ws.merge_cells(start_row=rlo + 1, end_row=rhi, start_column=clo + 1, end_column=chi)
                except Exception: pass
        wb_out.save(out_xlsx)
    except Exception:
        try:
            if os.path.isfile(out_xlsx): os.remove(out_xlsx)
        except OSError: pass
        return False
    return is_real_xlsx(out_xlsx)

def find_libreoffice_soffice():
    candidates = [os.environ.get("SOFFICE_PATH"), r"C:\Program Files\LibreOffice\program\soffice.exe", r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"]
    for p in candidates:
        if p and os.path.isfile(p): return p
    for name in ("soffice", "libreoffice"):
        path = which(name)
        if path: return path
    return None

def convert_xls_to_xlsx_libreoffice(xls_path: str, out_dir: str) -> str | None:
    soffice = find_libreoffice_soffice()
    if not soffice: return None
    os.makedirs(out_dir, exist_ok=True)
    try:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", os.path.abspath(xls_path), "--outdir", os.path.abspath(out_dir)],
            check=True,
            capture_output=True,
            text=True,
            timeout=120
        )
    except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
        return None
    base = os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx"
    out = os.path.join(out_dir, base)
    return out if os.path.isfile(out) and is_real_xlsx(out) else None

def convert_xls_to_xlsx_excel_com(xls_path: str, out_xlsx: str) -> bool:
    if sys.platform != "win32": return False
    try: import win32com.client
    except ImportError: return False
    xls_path = os.path.abspath(xls_path)
    out_xlsx = os.path.abspath(out_xlsx)
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xls_path)
        wb.SaveAs(out_xlsx, FileFormat=51)
        wb.Close(False)
        return is_real_xlsx(out_xlsx)
    except Exception:
        return False
    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass

def convert_xls_to_xlsx_pyexcel(xls_path: str, out_xlsx: str) -> bool:
    try: import pyexcel as pe
    except ImportError: return False
    xls_path = os.path.abspath(xls_path)
    out_xlsx = os.path.abspath(out_xlsx)
    try:
        pe.save_book_as(file_name=xls_path, dest_file_name=out_xlsx)
        return is_real_xlsx(out_xlsx)
    except Exception: return False

def ensure_xlsx_from_xls(xls_path: str, out_xlsx: str) -> bool:
    if is_real_xlsx(out_xlsx): return True
    if os.path.isfile(out_xlsx) and not is_real_xlsx(out_xlsx):
        try: os.remove(out_xlsx)
        except OSError: pass
    out_dir = os.path.dirname(out_xlsx)
    if convert_xls_to_xlsx_excel_com(xls_path, out_xlsx): return True
    conv = convert_xls_to_xlsx_libreoffice(xls_path, out_dir)
    if conv:
        if os.path.abspath(conv) != os.path.abspath(out_xlsx):
            import shutil
            shutil.copy2(conv, out_xlsx)
        if is_real_xlsx(out_xlsx): return True
        try: os.remove(out_xlsx)
        except OSError: pass
    if convert_xls_to_xlsx_xlrd_openpyxl(xls_path, out_xlsx): return True
    if convert_xls_to_xlsx_pyexcel(xls_path, out_xlsx): return True
    return False

def read_excel_any(path: str, tmp_dir: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv": return pd.read_csv(path, header=None, encoding="latin-1")
    if ext == ".xlsx": return pd.read_excel(path, header=None, engine="openpyxl")
    if ext == ".xls":
        try: return pd.read_excel(path, header=None, engine="xlrd")
        except Exception:
            conv = convert_xls_to_xlsx_libreoffice(path, tmp_dir)
            if conv and is_real_xlsx(conv): return pd.read_excel(conv, header=None, engine="openpyxl")
            base = os.path.splitext(os.path.basename(path))[0] + ".conv.xlsx"
            tmp_xlsx = os.path.join(tmp_dir, base)
            if convert_xls_to_xlsx_xlrd_openpyxl(path, tmp_xlsx): return pd.read_excel(tmp_xlsx, header=None, engine="openpyxl")
            raise
    return pd.read_excel(path, header=None)

def wait_for_folder_access(folder_path: str, retries: int = 3, delay_seconds: int = 2) -> bool:
    for _ in range(retries):
        if os.path.isdir(folder_path):
            try:
                os.listdir(folder_path)
                return True
            except OSError:
                time.sleep(delay_seconds)
        else:
            time.sleep(delay_seconds)
    return False

def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    try:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    except Exception:
        pass
    for col_idx in range(1, max_col + 1):
        src_cell = ws.cell(row=source_row, column=col_idx)
        dst_cell = ws.cell(row=target_row, column=col_idx)
        if isinstance(dst_cell, MergedCell):
            continue
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)

def find_start_row_excel_com(ws) -> int:
    # Tim dong ngay sau tieu de "STT" de do du lieu.
    max_row = min(int(getattr(ws.UsedRange, "Rows", ws.UsedRange).Count or 200) + 20, 500)
    max_col = min(int(getattr(ws.UsedRange, "Columns", ws.UsedRange).Count or 30) + 5, 60)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.Cells(r, c).Value
            if v is not None and str(v).strip() == "STT":
                return r + 1
    return 9

def write_combined_with_excel_com(template_path: str, output_file: str, combined: pd.DataFrame) -> bool:
    if sys.platform != "win32":
        return False
    try:
        import win32com.client
    except ImportError:
        return False

    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(template_path))
        ws = wb.Worksheets(1)
        start_row = find_start_row_excel_com(ws)
        for i, row in combined.iterrows():
            target_row = start_row + i
            for j, value in enumerate(row):
                ws.Cells(target_row, j + 1).Value = "" if pd.isna(value) else value
        wb.SaveAs(os.path.abspath(output_file), FileFormat=51)
        return True
    except Exception:
        return False
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass

def process_excel_in_folder(folder_path: str, template_filename: str) -> dict:
    folder_path = os.path.abspath(folder_path)
    template_name_input = template_filename

    if not wait_for_folder_access(folder_path):
        raise HTTPException(status_code=400, detail=f"Khong tim thay thu muc: {folder_path}")

    base_name, input_ext = os.path.splitext(template_name_input)
    input_ext = input_ext.lower()
    template_file = os.path.join(folder_path, f"{base_name}.xlsx")
    tpl_xls = os.path.join(folder_path, f"{base_name}.xls")
    output_file = os.path.join(folder_path, "Ket_qua_Tong_Hop.xlsx")
    tmp_dir = os.path.join(folder_path, "_rod_tmp_convert")
    os.makedirs(tmp_dir, exist_ok=True)

    template_source_for_com = None
    if input_ext == ".xls" and os.path.isfile(tpl_xls):
        template_source_for_com = tpl_xls
    elif os.path.isfile(template_file):
        template_source_for_com = template_file

    patterns = [os.path.join(folder_path, "*.xls*"), os.path.join(folder_path, "*.csv")]
    all_files = []
    for p in patterns:
        all_files.extend(glob.glob(p))
    all_files = [f for f in sorted(set(all_files)) if base_name not in f and "Ket_qua_Tong_Hop" not in f]

    all_data = []
    file_errors = []
    for file in all_files:
        try:
            df = read_excel_any(file, tmp_dir)
            data_part = df.iloc[2:, :10].copy()
            data_part.iloc[:, 2] = data_part.iloc[:, 2].apply(excel_date_to_str)
            data_part = remove_footer(data_part)
            all_data.append(data_part)
        except Exception as e:
            file_errors.append({"file": file, "error": str(e)})

    if not all_data:
        return {"status": "warning", "message": "Khong co du lieu gop.", "file_errors": file_errors}

    combined = pd.concat(all_data, ignore_index=True)
    # Uu tien ghi bang Excel COM de giu logo/shape/merge/format goc.
    if template_source_for_com and write_combined_with_excel_com(template_source_for_com, output_file, combined):
        return {"status": "success", "output_file": output_file, "file_errors": file_errors, "writer": "excel_com"}

    if input_ext == ".xls" and os.path.isfile(tpl_xls):
        try:
            if os.path.isfile(template_file):
                os.remove(template_file)
        except OSError:
            pass
        ensure_xlsx_from_xls(tpl_xls, template_file)
    elif not os.path.isfile(template_file) and os.path.isfile(tpl_xls):
        ensure_xlsx_from_xls(tpl_xls, template_file)

    if not os.path.isfile(template_file):
        raise HTTPException(status_code=404, detail=f"Thieu file Template: {base_name}")

    wb = load_workbook(template_file)
    ws = wb.active
    start_row = find_start_row(ws)
    style_source_row = start_row
    for i, row in combined.iterrows():
        target_row = start_row + i
        copy_row_style(ws, style_source_row, target_row, 10)
        for j, value in enumerate(row):
            cell = ws.cell(row=target_row, column=j + 1)
            if not isinstance(cell, MergedCell):
                cell.value = value

    wb.save(output_file)
    return {"status": "success", "output_file": output_file, "file_errors": file_errors, "writer": "openpyxl"}

# --- 3. ENDPOINT XỬ LÝ CHÍNH TỪ N8N ---
# ==================================================
# 3. ENDPOINT XỬ LÝ (Áp dụng Template động)
# ==================================================
# ==================================================
# 3. ENDPOINT XỬ LÝ CHÍNH
# ==================================================
@app.post("/process-excel")
def process_excel(req: MergeRequest):
    if req.run_in_background:
        job_id = str(uuid.uuid4())
        JOBS[job_id] = {
            "status": "queued",
            "created_at": datetime.now().isoformat(),
            "input": {
                "folder_path": req.folder_path,
                "template_filename": req.template_filename
            }
        }
        # Chay job tren thread rieng de request tra ve ngay va API van dap ung poll status.
        worker = threading.Thread(target=run_excel_job, args=(job_id, req), daemon=True)
        worker.start()
        return {
            "status": "accepted",
            "job_id": job_id,
            "check_status_at": f"/jobs/{job_id}",
            "message": "Da dua vao hang doi, vui long goi endpoint trang thai de lay ket qua."
        }
    return _process_excel_internal(req)

# ==================================================
# ENDPOINT UPLOAD NHIỀU FILE - ĐÃ SỬA
# ==================================================
# --- ĐỊNH NGHĨA CẤU TRÚC JSON NHẬN TỪ N8N ---
class FileData(BaseModel):
    filename: str
    content: str  # Chuỗi Base64 của file

class UploadRequest(BaseModel):
    output_filename: str = "Ket_qua_Tong_Hop.xlsx"
    template_file: FileData
    data_files: List[FileData]

# --- ENDPOINT XỬ LÝ ---
@app.post("/process-excel-upload", dependencies=[Depends(verify_api_key)])
async def process_excel_upload(
    request: UploadRequest,
    background_tasks: BackgroundTasks
):
    """
    API nhận JSON chứa dữ liệu file đã mã hóa Base64:
    - template_file: { filename: "...", content: "base64_string" }
    - data_files: [{ filename: "...", content: "base64_string" }, ...]
    - output_filename: tên file kết quả
    """
    
    if not request.template_file:
        raise HTTPException(status_code=400, detail="Thiếu template_file")
    if not request.data_files or len(request.data_files) == 0:
        raise HTTPException(status_code=400, detail="Cần ít nhất 1 file dữ liệu")
        
    if not request.template_file.filename.lower().endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="template_file phải là .xls hoặc .xlsx")

    # ===== TẠO THƯ MỤC TẠM =====
    work_dir = tempfile.mkdtemp(prefix="n8n_excel_")
    background_tasks.add_task(shutil.rmtree, work_dir, ignore_errors=True)

    try:
        # ===== GIẢI MÃ VÀ LƯU TEMPLATE =====
        template_name = os.path.basename(request.template_file.filename)
        template_path = os.path.join(work_dir, template_name)

        with open(template_path, "wb") as f:
            f.write(base64.b64decode(request.template_file.content))

        # ===== GIẢI MÃ VÀ LƯU DATA FILES =====
        for idx, file_data in enumerate(request.data_files, start=1):
            filename = os.path.basename(file_data.filename or f"file_{idx}.xlsx")

            if not filename.lower().endswith((".xls", ".xlsx", ".csv")):
                continue

            file_path = os.path.join(work_dir, f"{idx:04d}_{filename}")

            with open(file_path, "wb") as f:
                f.write(base64.b64decode(file_data.content))

        # ===== XỬ LÝ MERGE =====
        # (Giữ nguyên logic gọi hàm xử lý của bạn)
        result = process_excel_in_folder(work_dir, template_name)

        if result.get("status") not in {"success", "warning"}:
            raise HTTPException(status_code=500, detail=result)

        output_file = result.get("output_file")

        if not output_file or not os.path.isfile(output_file):
            raise HTTPException(status_code=500, detail="Không tạo được file kết quả")

        # ===== TRẢ VỀ FILE =====
        safe_name = os.path.basename(request.output_filename)
        if not safe_name.lower().endswith(".xlsx"):
            safe_name += ".xlsx"

        return FileResponse(
            output_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=safe_name
        )

    except Exception as e:
        shutil.rmtree(work_dir, ignore_errors=True)
        raise e


def _process_excel_internal(req: MergeRequest):
    return process_excel_in_folder(req.folder_path, req.template_filename)

def run_excel_job(job_id: str, req: MergeRequest):
    JOBS[job_id]["status"] = "running"
    JOBS[job_id]["started_at"] = datetime.now().isoformat()
    try:
        result = _process_excel_internal(req)
        JOBS[job_id]["status"] = result.get("status", "success")
        JOBS[job_id]["result"] = result
    except HTTPException as e:
        JOBS[job_id]["status"] = "failed"
        JOBS[job_id]["error"] = {
            "code": e.status_code,
            "detail": e.detail
        }
    except Exception:
        JOBS[job_id]["status"] = "failed"
        JOBS[job_id]["error"] = {
            "code": 500,
            "detail": traceback.format_exc()
        }
    finally:
        JOBS[job_id]["finished_at"] = datetime.now().isoformat()

@app.get("/jobs/{job_id}")
def get_job_status(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Khong tim thay job_id")
    return job

@app.post("/debug-path")
def debug_path(req: PathDebugRequest):
    folder_path = os.path.abspath(req.folder_path)
    result = {
        "folder_path": folder_path,
        "exists": os.path.isdir(folder_path),
        "listable": False,
        "read_ok": False,
        "write_ok": False,
        "sample_files": [],
        "error": None
    }
    if not os.path.isdir(folder_path):
        return result
    try:
        names = os.listdir(folder_path)
        result["listable"] = True
        result["sample_files"] = names[:20]
        # Thu doc thuoc tinh file dau tien neu co.
        if names:
            first_path = os.path.join(folder_path, names[0])
            os.path.exists(first_path)
        result["read_ok"] = True
    except Exception as e:
        result["error"] = f"READ/LIST failed: {e}"
        return result
    test_file = os.path.join(folder_path, f".api_write_test_{uuid.uuid4().hex}.tmp")
    try:
        with open(test_file, "w", encoding="utf-8") as f:
            f.write("ok")
        result["write_ok"] = True
    except Exception as e:
        result["error"] = f"WRITE failed: {e}"
    finally:
        try:
            if os.path.isfile(test_file):
                os.remove(test_file)
        except OSError:
            pass
    return result

@app.get("/")
def root():
    return {"message": "API is running. Connect via LAN IP if using other computers."}

# ==================================================
# 4. KHỞI CHẠY (QUAN TRỌNG ĐỂ CHẠY LAN)
# ==================================================
def main():
 
    # Tự động mở Firewall port 8000
    open_firewall_port(8000)
    
    # Lấy IP trong mạng LAN để hiển thị cho người dùng biết
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    
    print(f"[*] API dang chay tai link noi bo: http://{local_ip}:8000")
    print(f"[*] Link Localhost: http://127.0.0.1:8000")

    print("----------------------------------------------")
    
    # Host="0.0.0.0" cho phép các máy khác trong mạng LAN truy cập vào
    uvicorn.run(app, host="0.0.0.0", port=8000)

if __name__ == "__main__":
    main()
